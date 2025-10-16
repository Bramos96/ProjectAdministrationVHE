import os
import glob
import pandas as pd
import logging

# ───────────────────────────────────────────────────────────
# PAS JE PADEN HIER AAN
INPUT_FOLDER  = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Input"
CENTRAL_FILE  = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Overzicht Projectadministratie.xlsx"
MAPPING_FILE  = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Kolommenmapping per bron.xlsx"
OUTPUT_FOLDER = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Output"
# ───────────────────────────────────────────────────────────

logging.basicConfig(level=logging.DEBUG, format="%(message)s")
log = logging.getLogger()

def normalize_idx(x):
    num = pd.to_numeric(x, errors="coerce")
    return str(int(num)) if pd.notna(num) and float(num).is_integer() else x

# Handmatige kolommen
HANDMATE_COLUMNS = [
    "Algemene informatie",
    "Actiepunten Overig",
    "Verwacht resultaat"
]

# 1) Mappingslijst
mappings = {
    "Projectoverzicht Sumatra": {
        "Project": "Projectnummer", "Omschrijving": "Omschrijving",
        "Projectleider": "Projectleider", "Klant": "Klant",
        "Einddatum": "Einddatum", "Bud.Kost.": "Budget Kosten",
        "Bud.Opbr.": "Budget Opbrengsten", "Kosten": "Werkelijke kosten",
        "Opbrengsten": "Werkelijke opbrengsten", "Selcode": "Selcode",
        "Volg.lev.dat.": "Leverdatum"
    },
    "Verkoopdummy Sumatra": {
        "Order": "Projectnummer",
        "Niet toegewezen regel(s)": "Niet toegewezen regels",
        "Niet toegewezen": "Niet toegewezen regels"
    },
    "Werkbestand Projectadministratie": {
        "Project": "Projectnummer", "Omschrijving": "Omschrijving",
        "Verwacht resultaat": "Verwacht resultaat",
        "Actiepunten projectleider": "Actiepunten Projectleider",
        "Actiepunten Bram": "Actiepunten Bram",
        "Actiepunten Overig": "Actiepunten Overig",
        "Algemene informatie": "Algemene informatie"
    },
    "Overzicht Projectadministratie": {
        "Projectnummer": "Projectnummer", "Omschrijving": "Omschrijving",
        "Projectleider": "Projectleider", "Klant": "Klant",
        "Einddatum": "Einddatum", "B. Kosten": "Budget Kosten",
        "B. Opbrengst": "Budget Opbrengten", "W. Kosten": "Werkelijke kosten",
        "W. Opbrengst": "Werkelijke opbrengsten", "Leverdatum": "Leverdatum",
        "Niet toegewezen": "Niet toegewezen regels",
        "Verwacht resultaat": "Verwacht resultaat",
        "Actiepunten Projectleider": "Actiepunten Projectleider",
        "Actiepunten Bram": "Actiepunten Bram",
        "Actiepunten Overig": "Actiepunten Overig",
        "Whitelist ": "Whitelist ", "Algemene informatie": "Algemene informatie",
        "Versielog ": "Versielog", "Type": "Type"
    }
}
# strip whitespace
for key in mappings:
    mappings[key] = {k.strip(): v.strip() for k, v in mappings[key].items()}

# 2) Lees & standaardiseer CENTRAL
df_central = pd.read_excel(CENTRAL_FILE, header=1)
df_central.columns = df_central.columns.str.strip()
df_central.rename(columns=mappings["Overzicht Projectadministratie"], inplace=True)
df_central.set_index("Projectnummer", inplace=True)
central_cols = df_central.columns.tolist()

log.debug(">> After reading CENTRAL:")
log.debug(df_central[HANDMATE_COLUMNS].head().to_string())

# ───────────────────────────────────────────────────────────
# ───────────────────────────────────────────────────────────
# 2a) Reload latest werkbestand → capture only STYLE‐highlighted overrides
import openpyxl

pattern = os.path.join(OUTPUT_FOLDER, "Werkbestand_AlleProjecten_*.xlsx")
wb_files = glob.glob(pattern)
override_set = set()

if wb_files:
    wb_files.sort(key=os.path.getmtime, reverse=True)
    latest_wb = wb_files[0]
    log.debug(f"Inspecting styles in Werkbestand: {latest_wb}")

    # Load with openpyxl to inspect fills
    wb_style = openpyxl.load_workbook(latest_wb, data_only=True)
    ws_style = wb_style.active  # or use wb_style["SheetName"] if not the first sheet

    # Find the column number for "Verwacht resultaat"
    header_row = list(ws_style.iter_rows(min_row=1, max_row=1, values_only=False))[0]
    vr_col = None
    for idx, cell in enumerate(header_row):
        if cell.value and str(cell.value).strip() == "Verwacht resultaat":
            vr_col = idx + 1
            break

    if vr_col is None:
        log.debug("Could not find 'Verwacht resultaat' header in Werkbestand.")
    else:
        # Now scan each data row for orange fill (ARGB 'FFFFD966')
        for row in ws_style.iter_rows(min_row=2, min_col=vr_col, max_col=vr_col):
            cell = row[0]
            color = getattr(cell.fill.fgColor, "rgb", None)
            if color == "FFFFD966":
                # projectnummer is in column A (col 1)
                proj = ws_style.cell(row=cell.row, column=1).value
                override_set.add(str(proj))

    # Finally, copy the manual columns into df_central as before
    df_wb = pd.read_excel(latest_wb, header=1)
    df_wb.columns = df_wb.columns.str.strip()
    df_wb.rename(columns=mappings["Werkbestand Projectadministratie"], inplace=True)
    df_wb.set_index("Projectnummer", inplace=True)
    for col in HANDMATE_COLUMNS:
        if col in df_wb.columns and col in df_central.columns:
            df_central[col] = df_wb[col]

else:
    log.debug("No Werkbestand found; manual fields left as-is.")
# ───────────────────────────────────────────────────────────



# 3) Input-cols uit mapping
mapping_df = pd.read_excel(MAPPING_FILE, header=0)
mapping_df.rename(columns={mapping_df.columns[0]: "Header"}, inplace=True)
mapping_df.columns = mapping_df.columns.str.strip()
mask = mapping_df["Soort"].str.strip().str.lower() == "input"
input_cols = mapping_df.loc[mask.fillna(False), "Header"].dropna().tolist()
if "Selcode" not in input_cols:
    input_cols.append("Selcode")

# 4) Vind 2 nieuwste .xlsx in INPUT
files = [f for f in os.listdir(INPUT_FOLDER)
         if f.lower().endswith(".xlsx") and not f.startswith("~$")]
files.sort(key=lambda f: os.path.getmtime(os.path.join(INPUT_FOLDER, f)),
           reverse=True)
latest = files[:2]

# 5) Verwerk SUMATRA bestanden
all_inputs = []
dummy_list = []
for fname in latest:
    fp = os.path.join(INPUT_FOLDER, fname)
    df = pd.read_excel(fp, header=1)
    df.columns = df.columns.str.strip()
    cols_l = df.columns.str.lower().str.strip()

    if "bud.kost." in cols_l and ("projectleider" in cols_l or "selcode" in cols_l):
        typ = "Projectoverzicht Sumatra"
    elif any("niet toegewezen" in c for c in cols_l):
        typ = "Verkoopdummy Sumatra"
    elif "verwacht resultaat" in cols_l and "actiepunten bram" in cols_l:
        typ = "Werkbestand Projectadministratie"
    elif "versielog" in cols_l:
        typ = "Overzicht Projectadministratie"
    else:
        continue

    sub = (
        df[[c for c in mappings[typ] if c in df.columns]]
        .rename(columns=mappings[typ])
    )
    sub.set_index("Projectnummer", inplace=True)
    sub.index = sub.index.map(normalize_idx)
    sub.index.name = "Projectnummer"

    if typ == "Verkoopdummy Sumatra":
        dummy_list.append(sub[["Niet toegewezen regels"]])
    else:
        sel = [c for c in input_cols if c != "Projectnummer"]
        for col in HANDMATE_COLUMNS:
            if col in sub.columns and col not in sel:
                sel.append(col)
        all_inputs.append(sub.reindex(columns=sel))

# 6) Combineer inputs
if all_inputs:
    df_input = pd.concat(all_inputs, axis=0)
else:
    df_input = pd.DataFrame(columns=input_cols + HANDMATE_COLUMNS)

# Bescherm manual cols
for col in HANDMATE_COLUMNS:
    df_input[col] = pd.NA

log.debug(">> INPUT manual cols snapshot:\n" +
          df_input[HANDMATE_COLUMNS].head(10).to_string())

# 7) Dummyregels
if dummy_list:
    df_dummy = pd.concat(dummy_list, axis=0)
    df_dummy.index = df_dummy.index.map(normalize_idx)
    df_dummy.index.name = "Projectnummer"
    df_input.update(df_dummy)

# 8) Merge
exist = df_input.index.intersection(df_central.index)
new   = df_input.index.difference(df_central.index)

cols_to_update = [c for c in df_input.columns if c not in HANDMATE_COLUMNS]
if not exist.empty and cols_to_update:
    df_central.loc[exist, cols_to_update] = df_input.loc[exist, cols_to_update]

for col in HANDMATE_COLUMNS:
    if col in df_input.columns and col in df_central.columns:
        for idx in exist:
            val = df_input.at[idx, col]
            if pd.notna(val) and str(val).strip():
                df_central.at[idx, col] = val

# 9) Voeg nieuwe projecten toe
df_merged = pd.concat([df_central, df_input.loc[new]]).reset_index()

# 10) Voeg Type toe
if "Selcode" in df_merged.columns:
    df_merged["Type"] = df_merged["Selcode"].apply(
        lambda x: "Production" if str(x).strip() == "Orders Kabelafdeling"
        else "Proto"
    )
    if "Type" not in central_cols:
        central_cols.append("Type")

# 11) Herorden
cols_final = ["Projectnummer"] + [c for c in central_cols if c in df_merged.columns]
df_merged = df_merged[cols_final]

# 12) Bereken Verwacht resultaat
if "Budget Opbrengsten" in df_merged.columns and "Budget Kosten" in df_merged.columns:
    mask = df_merged["Verwacht resultaat"].isna() | (df_merged["Verwacht resultaat"] == "")
    df_merged.loc[mask, "Verwacht resultaat"] = (
        df_merged.loc[mask, "Budget Opbrengsten"]
        - df_merged.loc[mask, "Budget Kosten"]
    )

# 13) Datumformatting
for col in ["Einddatum", "Eerstvolgende leverdatum", "Leverdatum"]:
    if col in df_merged:
        df_merged[col] = (
            pd.to_datetime(df_merged[col], errors="coerce")
              .dt.strftime("%Y-%m-%d")
        )

# 14) Versielog
if "Versielog" in df_merged.columns:
    df_merged["Versielog"] = pd.Timestamp.today().strftime("%Y-%m-%d")

# 15) Valutavelden
currency_cols = [
    "Budget Kosten","Budget Opbrengsten",
    "Werkelijke kosten","Werkelijke opbrengsten","Verwacht resultaat"
]
for col in currency_cols:
    if col in df_merged:
        df_merged[col] = df_merged[col].fillna(0).round(0).astype(int)

# 16) Export + highlight overrides
dt = pd.Timestamp.today()
fn = f"Overzicht_Projectadministratie_Week{dt.week}_{dt.year}.xlsx"
out_path = os.path.join(OUTPUT_FOLDER, fn)

with pd.ExcelWriter(out_path, engine="xlsxwriter") as writer:
    df_merged.to_excel(writer, index=False, sheet_name="Overzicht")
    wb = writer.book
    ws = writer.sheets["Overzicht"]

    # formats
    money_fmt  = wb.add_format({"num_format":"€#,##0"})
    header_fmt = wb.add_format({
        "bold":True,"font_color":"white","bg_color":"#4F81BD",
        "align":"center","valign":"vcenter","font_size":12
    })
    orange_fmt = wb.add_format({"bg_color":"#FFD966"})

    # write headers
    for i, col in enumerate(df_merged.columns):
        ws.write(0, i, col, header_fmt)
    ws.set_row(0, 25)
    max_row, max_col = df_merged.shape
    ws.autofilter(0, 0, max_row, max_col-1)
    ws.freeze_panes(1, 0)

    # adjust widths and apply money fmt
    for i, col in enumerate(df_merged.columns):
        width = max(df_merged[col].astype(str).map(len).max(), len(col)) + 2
        fmt = money_fmt if col in currency_cols else None
        ws.set_column(i, i, width, fmt)

    # highlight manual overrides in "Verwacht resultaat"
    col_idx = df_merged.columns.get_loc("Verwacht resultaat")
    for row_idx, proj in enumerate(df_merged["Projectnummer"], start=1):
        if str(proj) in override_set:
            val = df_merged.at[row_idx-1, "Verwacht resultaat"]
            ws.write(row_idx, col_idx, val, orange_fmt)

print(f"Klaar, opgeslagen als: {fn}")
