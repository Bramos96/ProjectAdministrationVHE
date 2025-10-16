import os
import pandas as pd
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# ───────────────────────────────────────────────────────────
# PAS JE PADEN HIER AAN: pad naar je week‐file
OUTPUT_FILE = (
    r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration"
    r"\Output\Overzicht_Projectadministratie_Week{week}_{year}.xlsx"
)
# ───────────────────────────────────────────────────────────

def make_actions(row, today):
    """Return a bullet‐list for one project row."""
    bullets = []

    ed_ts = pd.to_datetime(row.get("Einddatum"), errors="coerce")
    if pd.notna(ed_ts) and today > ed_ts.date():
        bullets.append("• Einddatum verlopen")

    lv_ts = pd.to_datetime(row.get("Eerstvolgende leverdatum"), errors="coerce")
    if pd.notna(lv_ts) and today > lv_ts.date():
        bullets.append("• Leverdatum(s) verlopen")

    return "\n".join(bullets)


def main():
    today = date.today()
    week  = today.isocalendar()[1]
    year  = today.year

    path = OUTPUT_FILE.format(week=week, year=year)
    if not os.path.exists(path):
        raise FileNotFoundError(f"Kun bestand niet vinden: {path}")

    # 1) Lees het overzicht met echte header
    df = pd.read_excel(path, header=0, engine="openpyxl")

    # 2) Bereken de nieuwe Actiepunten-kolom
    df["Actiepunten Projectleider"] = df.apply(lambda r: make_actions(r, today), axis=1)

    # 3) Laad werkboek & vind kolom
    wb = load_workbook(path)
    ws = wb["Overzicht"]
    col_letter = None
    for cell in ws[1]:
        if cell.value == "Actiepunten Projectleider":
            col_letter = cell.column_letter
            break
    if col_letter is None:
        raise ValueError("Kolom 'Actiepunten Projectleider' niet gevonden")

    # 4) Overschrijf per project, met wrap_text en aangepaste hoogte
    updated = 0
    for idx, text in enumerate(df["Actiepunten Projectleider"], start=2):
        cell = ws[f"{col_letter}{idx}"]
        if text:
            cell.value = text
            cell.alignment = Alignment(wrap_text=True)
            # pas rijhoogte aan op basis van aantal regels
            lines = text.count("\n") + 1
            ws.row_dimensions[idx].height = lines * 15
            updated += 1

    wb.save(path)
    print(f"✅ {updated} rijen bijgewerkt in kolom 'Actiepunten Projectleider' van {os.path.basename(path)}")

if __name__ == "__main__":
    main()
