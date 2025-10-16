import os
import pandas as pd
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# ───────────────────────────────────────────────────────────
# PAS JE PADEN HIER AAN
OUTPUT_FILE = (
    r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration"
    r"\Output\Overzicht_Projectadministratie_Week{week}_{year}.xlsx"
)
# ───────────────────────────────────────────────────────────

def check_klaar_om_te_sluiten(row):
    """
    Bepaal of project klaar is om te sluiten.
    Return True/False.
    """
    bo_raw = row.get("Budget Opbrengsten")
    wo_raw = row.get("Werkelijke opbrengsten")

    try:
        bo = float(str(bo_raw).replace(".", "").replace(",", ".").strip()) if pd.notna(bo_raw) else None
        wo = float(str(wo_raw).replace(".", "").replace(",", ".").strip()) if pd.notna(wo_raw) else None

        if bo is not None and wo is not None:
            if bo != 0 and bo <= wo:
                return True
    except Exception as e:
        print(f"⚠️ Fout bij check_klaar_om_te_sluiten: {e}")

    return False

def make_actions(row, today, klaar_om_te_sluiten):
    """Return a bullet‐list for Actiepunten Projectleider."""
    bullets = []

    # 1) Einddatum verlopen?
    if not klaar_om_te_sluiten:
        ed_ts = pd.to_datetime(row.get("Einddatum"), errors="coerce")
        if pd.notna(ed_ts) and today > ed_ts.date():
            bullets.append("• Einddatum verlopen")

    # 2) Leverdatum(s) verlopen?
    lv_ts = pd.to_datetime(row.get("Leverdatum"), errors="coerce")
    if pd.notna(lv_ts) and today > lv_ts.date():
        bullets.append("• Leverdatum(s) verlopen")

    # 3) Budget Kosten ontbreekt?
    bk = row.get("Budget Kosten")
    try:
        if pd.notna(bk) and float(bk) == 0:
            bullets.append("• Budget kosten toevoegen")
    except:
        pass

    # 4) Budget Opbrengsten ontbreekt?
    bo = row.get("Budget Opbrengsten")
    try:
        if pd.notna(bo) and float(bo) == 0:
            bullets.append("• Budget opbrengsten toevoegen")
    except:
        pass

    return "\n".join(bullets)

def make_actions_bram(row, klaar_om_te_sluiten):
    """Return a bullet‐list for Actiepunten Bram."""
    bullets = []

    # (1) Negatief Verwacht resultaat
    vr = row.get("Verwacht resultaat")
    try:
        if pd.notna(vr) and float(vr) < 0:
            bullets.append("• Negatief resultaat bespreken")
    except:
        pass

    # (2) Klaar om te sluiten?
    if klaar_om_te_sluiten:
        bullets.append("• Klaar om te sluiten?")

    return "\n".join(bullets)

def main():
    today = date.today()
    week  = today.isocalendar()[1]
    year  = today.year

    path = OUTPUT_FILE.format(week=week, year=year)
    if not os.path.exists(path):
        raise FileNotFoundError(f"Kon bestand niet vinden: {path}")

    # 1) Lees bestand
    df = pd.read_excel(path, header=0, engine="openpyxl")

    # Voeg twee nieuwe kolommen toe
    df["Actiepunten Projectleider"] = ""
    df["Actiepunten Bram"] = ""

    # 2) Loop door rijen
    for idx, row in df.iterrows():
        klaar = check_klaar_om_te_sluiten(row)

        ap_proj = make_actions(row, today, klaar)
        ap_bram = make_actions_bram(row, klaar)

        # → Check of beide conclusies aanwezig zijn:
        leverdatum_text = "• Leverdatum(s) verlopen"
        klaar_text = "• Klaar om te sluiten?"

        # Detectie of beide conclusies er zijn
        if leverdatum_text in ap_proj and klaar_text in ap_bram:
            # Verwijder beide conclusies
            ap_proj = "\n".join(
                [line for line in ap_proj.split("\n") if line != leverdatum_text]
            )
            ap_bram = "\n".join(
                [line for line in ap_bram.split("\n") if line != klaar_text]
            )

            # Voeg nieuwe conclusie toe
            ap_bram = "• Volledig gefactureerd, maar niet volledig geleverd?"

        df.at[idx, "Actiepunten Projectleider"] = ap_proj
        df.at[idx, "Actiepunten Bram"] = ap_bram

    # 3) Laad werkboek
    wb = load_workbook(path)
    ws = wb["Overzicht"]

    # Vind kolomletters
    col_letters = {}
    for cell in ws[1]:
        if cell.value in ["Actiepunten Projectleider", "Actiepunten Bram"]:
            col_letters[cell.value] = cell.column_letter

    # Voeg kolommen toe indien nodig
    for col in ["Actiepunten Projectleider", "Actiepunten Bram"]:
        if col not in col_letters:
            last_col = ws.max_column + 1
            ws.cell(row=1, column=last_col).value = col
            col_letters[col] = ws.cell(row=1, column=last_col).column_letter

    # 4) Overschrijf waarden
    for col_name in ["Actiepunten Projectleider", "Actiepunten Bram"]:
        col_letter = col_letters[col_name]

        for idx, text in enumerate(df[col_name], start=2):
            cell = ws[f"{col_letter}{idx}"]
            cell.value = text if text else None
            cell.alignment = Alignment(wrap_text=True)
            lines = text.count("\n") + 1 if text else 1
            ws.row_dimensions[idx].height = lines * 15

        ws.column_dimensions[col_letter].width = 50

    wb.save(path)
    print(f"✅ Script afgerond. Kolommen bijgewerkt in {os.path.basename(path)}")

if __name__ == "__main__":
    main()
