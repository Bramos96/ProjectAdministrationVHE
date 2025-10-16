import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

# ───────────────────────────────────────────────────────────
OUTPUT_FOLDER = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Output"
SYNC_COLUMNS = ["Algemene informatie", "Verwacht resultaat", "Actiepunten Overig"]
SHEET_NAME = "Overzicht"
ORANJE_FONT = Font(color="FF6600")
# ───────────────────────────────────────────────────────────

def find_latest_file(folder, prefix):
    files = [f for f in os.listdir(folder) if f.startswith(prefix) and f.lower().endswith(".xlsx")]
    if not files:
        raise FileNotFoundError(f"Geen bestand gevonden met prefix '{prefix}' in {folder}")
    files.sort(key=lambda x: os.path.getmtime(os.path.join(folder, x)), reverse=True)
    return os.path.join(folder, files[0])

# 1. Laad werkbestand (header op rij 2)
workfile_path = find_latest_file(OUTPUT_FOLDER, "Werkbestand_AlleProjecten")
df_work = pd.read_excel(workfile_path, sheet_name="Sheet1", header=1, engine="openpyxl")
df_work.columns = df_work.columns.str.strip().str.lower()

# Kolomnamen consistent maken
column_mapping = {
    "project": "projectnummer",
    "actiepunten projectleider": "actiepunten projectleider",
    "actiepunten bram": "actiepunten bram",
    "actiepunten overig": "actiepunten overig",
    "verwacht resultaat": "verwacht resultaat",
    "algemene informatie": "algemene informatie"
}
df_work.rename(columns=column_mapping, inplace=True)

# Zorg dat projectnummer string is
if "projectnummer" in df_work.columns:
    df_work["projectnummer"] = df_work["projectnummer"].astype(str).str.strip()
    df_work.set_index("projectnummer", inplace=True)

# 2. Laad originele overzichtbestand (header op rij 1)
overview_path = find_latest_file(OUTPUT_FOLDER, "Overzicht_Projectadministratie_Week")
df_overview = pd.read_excel(overview_path, sheet_name=SHEET_NAME, header=0, engine="openpyxl")
df_overview.columns = df_overview.columns.str.strip()
df_overview["Projectnummer"] = df_overview["Projectnummer"].astype(str).str.strip()

# 3. Open werkboek via openpyxl
wb = load_workbook(overview_path)
ws = wb[SHEET_NAME]

# 4. Zoek kolomletters in Excel
col_map = {}
for cell in ws[1]:
    if cell.value and cell.value.strip() in SYNC_COLUMNS:
        col_map[cell.value.strip()] = get_column_letter(cell.column)

# 5. Per rij controleren en bijwerken
updates = 0
for row_idx in range(2, ws.max_row + 1):
    projectnummer = ws[f"A{row_idx}"].value
    if projectnummer:
        projectnummer = str(projectnummer).strip()
        if projectnummer in df_work.index:
            for kolom in SYNC_COLUMNS:
                kolom_lower = kolom.lower()
                # Check of kolom überhaupt voorkomt in werkbestand
                if kolom_lower in df_work.columns:
                    cell_ref = f"{col_map[kolom]}{row_idx}"
                    new_val = df_work.at[projectnummer, kolom_lower]
                    old_val = ws[cell_ref].value

                    # Als werkbestand leeg is, niks overschrijven
                    if pd.notna(new_val) and str(new_val).strip() != "":
                        if str(new_val) != str(old_val):
                            ws[cell_ref].value = new_val
                            # Alleen tekstkleur oranje voor Verwacht resultaat
                            if kolom == "Verwacht resultaat":
                                ws[cell_ref].font = ORANJE_FONT
                            updates += 1

# 6. Overschrijf bestaand bestand
wb.save(overview_path)
print(f"Bestand overschreven: {overview_path}")
print(f"Gewijzigde celwaarden: {updates}")
