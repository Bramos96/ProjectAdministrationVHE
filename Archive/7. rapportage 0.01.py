import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ───────────────────────────────────────────────────────────
INPUT_FILE = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Output\Overzicht_Projectadministratie_Week26_2025.xlsx"
POWERBI_FILE = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\PowerBI.xlsx"
SHEET_NAME = "Overzicht"
# ───────────────────────────────────────────────────────────

# 1. Lees overzichtbestand
df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME, header=0)
df.columns = df.columns.str.strip()

# 2. Vul lege actiepunten met lege string
cols = ["Actiepunten Projectleider", "Actiepunten Bram", "Actiepunten Overig"]
for col in cols:
    if col in df.columns:
        df[col] = df[col].fillna("").astype(str)
    else:
        df[col] = ""

# 3. Combineer actiepunten
df["Actiepunten"] = df[cols].apply(lambda row: "; ".join([v.strip() for v in row if v.strip()]), axis=1)

# 4. Splits actiepunten naar rijen
records = []
for _, row in df.iterrows():
    project = row["Projectnummer"]
    actiepunten = row["Actiepunten"]
    if actiepunten:
        actiepunten_clean = actiepunten.replace("•", ";")
        for a in actiepunten_clean.split(";"):
            cleaned = a.strip().lstrip("•").strip()
            if cleaned:
                records.append({
                    "Projectnummer": project,
                    "Actiepunt": cleaned
                })

df_out = pd.DataFrame(records)

# 5. Schrijf naar Excel en stel kolombreedtes in
df_out.to_excel(POWERBI_FILE, index=False)

# Kolombreedte automatisch aanpassen
wb = load_workbook(POWERBI_FILE)
ws = wb.active

for col_idx, column_cells in enumerate(ws.columns, 1):
    max_length = max(len(str(cell.value or "")) for cell in column_cells)
    adjusted_width = max_length + 2
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = adjusted_width

wb.save(POWERBI_FILE)
print(f"✅ PowerBI.xlsx succesvol gevuld en netjes opgemaakt.")
