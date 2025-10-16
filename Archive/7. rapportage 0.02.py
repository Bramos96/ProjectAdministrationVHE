import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# ───────────────────────────────────────────────────────────
INPUT_FILE = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Output\Overzicht_Projectadministratie_Week26_2025.xlsx"
POWERBI_FILE = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\PowerBI.xlsx"
LOG_SHEET = "Log"
# ───────────────────────────────────────────────────────────

# 1. Week & jaar bepalen
today = datetime.today()
week = today.isocalendar()[1]
jaar = today.year

# 2. Lees overzichtbestand
df = pd.read_excel(INPUT_FILE, sheet_name="Overzicht", header=0)
df.columns = df.columns.str.strip()

# 3. Combineer en split actiepunten
cols = ["Actiepunten Projectleider", "Actiepunten Bram", "Actiepunten Overig"]
for col in cols:
    if col in df.columns:
        df[col] = df[col].fillna("").astype(str)
    else:
        df[col] = ""

df["Actiepunten"] = df[cols].apply(lambda row: "; ".join([v.strip() for v in row if v.strip()]), axis=1)

records = []
for _, row in df.iterrows():
    project = row["Projectnummer"]
    actiepunten = row["Actiepunten"]
    if actiepunten:
        actiepunten_clean = actiepunten.replace("•", ";")
        for a in actiepunten_clean.split(";"):
            cleaned = a.strip().lstrip("•").strip()
            if cleaned:
                records.append({
                    "Projectnummer": project,
                    "Actiepunt": cleaned,
                    "Week": week,
                    "Jaar": jaar
                })

df_new = pd.DataFrame(records)

# 4. Open of maak werkboek
if os.path.exists(POWERBI_FILE):
    wb = load_workbook(POWERBI_FILE)
else:
    wb = load_workbook(filename=None)

# 5. Haal of maak logsheet
if LOG_SHEET in wb.sheetnames:
    ws = wb[LOG_SHEET]

    # Huidige inhoud ophalen als DataFrame
    existing = pd.DataFrame(ws.values)
    existing.columns = existing.iloc[0]
    existing = existing.drop(0)

    # Filter bestaande data op andere weken dan huidige
    existing = existing[~((existing["Week"] == str(week)) & (existing["Jaar"] == str(jaar)))]
    df_combined = pd.concat([existing, df_new], ignore_index=True)
else:
    wb.create_sheet(LOG_SHEET)
    df_combined = df_new

# 6. Overschrijf logsheet
if LOG_SHEET in wb.sheetnames:
    ws = wb[LOG_SHEET]
    wb.remove(ws)
wb.create_sheet(LOG_SHEET)
ws = wb[LOG_SHEET]

for r in dataframe_to_rows(df_combined, index=False, header=True):
    ws.append(r)

# 7. Opslaan
wb.save(POWERBI_FILE)
print(f"✅ {len(df_new)} actiepunten gelogd voor week {week}, jaar {jaar}.")
