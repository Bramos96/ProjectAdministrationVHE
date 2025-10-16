import os
import pandas as pd
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# ───────────────────────────────────────────────────────────
# PAS JE PADEN HIER AAN: pad naar je week‐file
OUTPUT_FILE = (
    r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration"
    r"\Output\Overzicht_Projectadministratie_Week{week}_{year}.xlsx"
)
# ───────────────────────────────────────────────────────────

def make_actions(row, today):
    """Return a bullet‐list for Actiepunten Projectleider."""
    bullets = []

    # 1) Einddatum verlopen?
    ed_ts = pd.to_datetime(row.get("Einddatum"), errors="coerce")
    if pd.notna(ed_ts) and today > ed_ts.date():
        bullets.append("• Einddatum verlopen")

    # 2) Leverdatum(s) verlopen?
    lv_ts = pd.to_datetime(row.get("Leverdatum"), errors="coerce")
    if pd.notna(lv_ts) and today > lv_ts.date():
        bullets.append("• Leverdatum(s) verlopen")

    # 3) Budget Kosten ontbreekt?
    bk = row.get("Budget Kosten")
    try:
        if pd.notna(bk) and float(bk) == 0:
            bullets.append("• Budget kosten toevoegen")
    except:
        pass

    # 4) Budget Opbrengsten ontbreekt?
    bo = row.get("Budget Opbrengsten")
    try:
        if pd.notna(bo) and float(bo) == 0:
            bullets.append("• Budget opbrengsten toevoegen")
    except:
        pass

    return "\n".join(bullets)

def make_actions_bram(row):
    """Return a bullet‐list for Actiepunten Bram."""
    bullets = []

    # (1) Negatief Verwacht resultaat
    vr = row.get("Verwacht resultaat")
    try:
        if pd.notna(vr) and float(vr) < 0:
            bullets.append("• Negatief resultaat bespreken")
    except:
        pass

    # (2) Klaar om te sluiten?
    bo_raw = row.get("Budget Opbrengsten")
    wo_raw = row.get("Werkelijke opbrengsten")

    try:
        # Zorg dat getallen netjes geconverteerd worden
        bo = float(str(bo_raw).replace(".", "").replace(",", ".").strip()) if pd.notna(bo_raw) else None
        wo = float(str(wo_raw).replace(".", "").replace(",", ".").strip()) if pd.notna(wo_raw) else None

        if bo is not None and wo is not None:
            if bo != 0 and bo <= wo:
                bullets.append("• Klaar om te sluiten?")
    except Exception as e:
        print(f"⚠️ Fout bij vergelijken BO en WO: {e}")

    return "\n".join(bullets)

def main():
    today = date.today()
    week  = today.isocalendar()[1]
    year  = today.year

    path = OUTPUT_FILE.format(week=week, year=year)
    if not os.path.exists(path):
        raise FileNotFoundError(f"Kon bestand niet vinden: {path}")

    # 1) Lees het overzicht met echte header
    df = pd.read_excel(path, header=0, engine="openpyxl")

    # 2) Bereken nieuwe kolommen
    df["Actiepunten Projectleider"] = df.apply(lambda r: make_actions(r, today), axis=1)
    df["Actiepunten Bram"] = df.apply(make_actions_bram, axis=1)

    # 3) Laad werkboek & vind kolomletters
    wb = load_workbook(path)
    ws = wb["Overzicht"]

    # Vind kolomletters
    col_letters = {}
    for cell in ws[1]:  # header row
        if cell.value in ["Actiepunten Projectleider", "Actiepunten Bram"]:
            col_letters[cell.value] = cell.column_letter

    # Voeg kolom toe als hij nog niet bestaat
    if "Actiepunten Bram" not in col_letters:
        # Plaats hem rechts naast Actiepunten Projectleider (of helemaal achteraan)
        last_col = ws.max_column + 1
        ws.cell(row=1, column=last_col).value = "Actiepunten Bram"
        col_letters["Actiepunten Bram"] = ws.cell(row=1, column=last_col).column_letter

    # 4) Overschrijf waarden + wrap_text + hoogte
    for col_name in ["Actiepunten Projectleider", "Actiepunten Bram"]:
        col_letter = col_letters[col_name]

        for idx, text in enumerate(df[col_name], start=2):
            cell = ws[f"{col_letter}{idx}"]
            # Overschrijf ALTIJD, ook met leeg (None)
            cell.value = text if text else None
            cell.alignment = Alignment(wrap_text=True)
            lines = text.count("\n") + 1 if text else 1
            ws.row_dimensions[idx].height = lines * 15

        # Kolombreedte iets ruimer maken
        if col_name == "Actiepunten Bram":
            ws.column_dimensions[col_letter].width = 50
        else:
            ws.column_dimensions[col_letter].width = 40

    wb.save(path)
    print(f"✅ Script afgerond. Kolommen bijgewerkt in {os.path.basename(path)}")

if __name__ == "__main__":
    main()
