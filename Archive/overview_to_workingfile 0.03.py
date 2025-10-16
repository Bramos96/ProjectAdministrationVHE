import os
import pandas as pd
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ───────────────────────────────────────────────────────────
# PAS JE PADEN HIER AAN
OUTPUT_FOLDER = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Output"
TEMPLATE_FILE = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Werkbestand Projectadministratie.xlsx"
MAPPING_FILE  = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Kolommenmapping per bron.xlsx"
# ───────────────────────────────────────────────────────────

def find_latest_overview(folder):
    files = [f for f in os.listdir(folder)
             if f.startswith("Overzicht_Projectadministratie_Week") and f.lower().endswith(".xlsx")]
    if not files:
        raise FileNotFoundError(f"No overview file found in {folder}")
    files.sort(key=lambda x: os.path.getmtime(os.path.join(folder, x)), reverse=True)
    return os.path.join(folder, files[0])

# 1) Load central overview into DataFrame
overview_path = find_latest_overview(OUTPUT_FOLDER)
df = pd.read_excel(overview_path, header=0, engine="openpyxl")

# 2) Load mapping and build header map
df_map = pd.read_excel(MAPPING_FILE, header=0)
df_map.rename(columns={df_map.columns[0]: "Header"}, inplace=True)
df_map.columns = df_map.columns.str.strip()
in_mask = df_map["Soort"].str.strip().str.lower() == "input"
df_in = df_map.loc[in_mask]
header_map = {
    std.strip(): werk.strip()
    for std, werk in zip(df_in["Header"], df_in["Werkbestand Projectadministratie"])
    if isinstance(werk, str) and werk.strip()
}
# always include these columns
for col in ["Projectleider", "Actiepunten Projectleider", "Actiepunten Bram"]:
    if col in df.columns:
        header_map[col] = col

# 3) Open template workbook
wb = load_workbook(TEMPLATE_FILE)
ws = wb.active
# remove existing table definitions if any
if hasattr(ws, '_tables'):
    ws._tables.clear()

# 4) Identify template header columns (row 2)
template_cols = {}
for cell in ws[2]:
    if cell.value and cell.value.strip() in header_map.values():
        template_cols[cell.value.strip()] = cell.column
if not template_cols:
    raise ValueError("No matching template columns found in row 2")
# determine full shading range
header_cols = [cell.column for cell in ws[2] if cell.value is not None]
min_col, max_col = min(header_cols), max(header_cols)

# 5) Define styles
def get_styles():
    hdr_color = '4F81BD'
    header_fill  = PatternFill(start_color=hdr_color, end_color=hdr_color, fill_type='solid')
    header_font  = Font(color='FFFFFF', bold=True, size=12)
    data_font    = Font(size=12)
    align_style  = Alignment(horizontal='left', vertical='top', wrap_text=True)
    even_fill    = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    odd_fill     = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    return header_fill, header_font, data_font, align_style, even_fill, odd_fill

header_fill, header_font, data_font, align_style, even_fill, odd_fill = get_styles()

# 6) Write data rows (starting at row 3)
start_row = 3
for idx, row in df.iterrows():
    r = start_row + idx
    # fill each column
    for std, werk in header_map.items():
        if std in df.columns and werk in template_cols:
            c = template_cols[werk]
            val = row[std]
            cell = ws.cell(row=r, column=c, value=val)
            # wrap text for actiepunten columns
            if werk in ["Actiepunten Projectleider", "Actiepunten Bram"] and isinstance(val, str):
                parts = [p.strip() for p in val.replace(';', '\n').split('\n') if p.strip()]
                cell.value = '\n'.join(parts)
            # italic for Projectnummer
            if werk == "Projectnummer":
                cell.font = Font(size=12, italic=True)
            else:
                cell.font = data_font
            cell.alignment = align_style
    # shading across full width
    fill = even_fill if idx % 2 == 0 else odd_fill
    for c in range(min_col, max_col+1):
        ws.cell(row=r, column=c).fill = fill
    # set row height
    ws.row_dimensions[r].height = 50

# 7) Style header rows (rows 1 and 2)
for row_idx in [1, 2]:
    for c in range(min_col, max_col+1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='left', vertical='center')

# 8) Insert Date in A1
today_str = date.today().strftime("%Y-%m-%d")
ws["A1"] = f"Datum: {today_str}"

# 9) Freeze panes below header
ws.freeze_panes = 'A3'

# 10) Auto-fit columns (special for actiepunten)
total_rows = start_row + len(df) - 1
for werk, col_idx in template_cols.items():
    col_letter = get_column_letter(col_idx)
    if werk in ["Actiepunten Projectleider", "Actiepunten Bram"]:
        # both 50 wide
        ws.column_dimensions[col_letter].width = 50
    else:
        lens = [len(str(ws[f"{col_letter}2"].value or ''))]
        for r in range(start_row, total_rows+1):
            lens.append(len(str(ws[f"{col_letter}{r}"].value or '')))
        width = max(lens) + 2
        ws.column_dimensions[col_letter].width = width

# 11) Save as new workfile
today_txt = date.today().strftime("%Y-%m-%d")
out_fn = f"Werkbestand_AlleProjecten_{today_txt}.xlsx"
out_path = os.path.join(OUTPUT_FOLDER, out_fn)
wb.save(out_path)
print(f" Werkbestand gegenereerd en gestyled: {out_path}")
