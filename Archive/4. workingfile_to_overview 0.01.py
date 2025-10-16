import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

# ───────────────────────────────────────────────────────────
OUTPUT_FOLDER = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Output"
SYNC_COLUMNS = ["Algemene informatie", "Verwacht resultaat", "Actiepunten Overig"]
SHEET_NAME = "Overzicht"
ORANJE_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
# ───────────────────────────────────────────────────────────

def find_latest_file(folder, prefix):
    files = [f for f in os.listdir(folder) if f.startswith(prefix) and f.lower().endswith(".xlsx")]
    if not files:
        raise FileNotFoundError(f"Geen bestand gevonden met prefix '{prefix}' in {folder}")
    files.sort(key=lambda x: os.path.getmtime(os.path.join(folder, x)), reverse=True)
    return os.path.join(folder, files[0])

# 1. Laad werkbestand (header op rij 2)
workfile_path = find_latest_file(OUTPUT_FOLDER, "Werkbestand_AlleProjecten")
df_work = pd.read_excel(workfile_path, sheet_name="Sheet1", header=1, engine="openpyxl")
df_work.columns = df_work.columns.str.strip()
if "Project" in df_work.columns:
    df_work.rename(columns={"Project": "Projectnummer"}, inplace=True)
df_work["Projectnummer"] = df_work["Projectnummer"].astype(str).str.strip()
df_work.set_index("Projectnummer", inplace=True)

# 2. Laad originele overzichtbestand (header op rij 1)
overview_path = find_latest_file(OUTPUT_FOLDER, "Overzicht_Projectadministratie_Week26")
df_overview = pd.read_excel(overview_path, sheet_name=SHEET_NAME, header=0, engine="openpyxl")
df_overview.columns = df_overview.columns.str.strip()
df_overview["Projectnummer"] = df_overview["Projectnummer"].astype(str).str.strip()

# 3. Open het werkboek via openpyxl
wb = load_workbook(overview_path)
ws = wb[SHEET_NAME]

# 4. Zoek kolomletters voor te synchroniseren kolommen
col_map = {}
for cell in ws[1]:
    if cell.value in SYNC_COLUMNS:
        col_map[cell.value] = get_column_letter(cell.column)

# 5. Per rij controleren en bijwerken
updates = 0
for row_idx in range(2, ws.max_row + 1):  # Excel is 1-based
    projectnummer = ws[f"A{row_idx}"].value
    if projectnummer:
        projectnummer = str(projectnummer).strip()
        if projectnummer in df_work.index:
            for kolom in SYNC_COLUMNS:
                if kolom in col_map and kolom in df_work.columns:
                    cell_ref = f"{col_map[kolom]}{row_idx}"
                    new_val = df_work.at[projectnummer, kolom]
                    old_val = ws[cell_ref].value
                    if pd.notna(new_val) and new_val != old_val:
                        ws[cell_ref].value = new_val
                        if kolom == "Verwacht resultaat":
                            ws[cell_ref].fill = ORANJE_FILL
                        updates += 1

# 6. Overschrijf bestaand bestand
wb.save(overview_path)
print(f"✅ Bestand overschreven: {overview_path}")
print(f"🟧 Gewijzigde celwaarden: {updates}")
