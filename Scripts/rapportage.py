import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import glob

# ───────────────────────────────────────────────────────────
OUTPUT_FOLDER = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Output"
POWERBI_FILE = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\PowerBI.xlsx"
LOG_SHEET = "Log"
# ───────────────────────────────────────────────────────────

# 1. Zoek laatste overzicht-bestand
pattern = os.path.join(
    OUTPUT_FOLDER,
    "Overzicht_Projectadministratie_Week*.xlsx"
)
files = glob.glob(pattern)

if files:
    files.sort(key=os.path.getmtime, reverse=True)
    INPUT_FILE = files[0]
    print(f"👉 Laatste bestand gevonden: {INPUT_FILE}")
else:
    raise FileNotFoundError("Geen overzichtsbestand gevonden in de Output-folder.")

# 2. Week & jaar bepalen
today = datetime.today()
week = today.isocalendar()[1]
jaar = today.year

# 3. Lees overzichtbestand
df = pd.read_excel(INPUT_FILE, sheet_name="Overzicht", header=0)
df.columns = df.columns.str.strip()

print("Kolommen in df:", df.columns.tolist())
print("Voorbeeld Projectleider waardes:", df["Projectleider"].head(10).tolist())
print("Voorbeeld Type waardes:", df["Type"].head(10).tolist())


# 4. Combineer en split actiepunten
cols = ["Actiepunten Projectleider", "Bespreekpunten", "Actiepunten Bram"]
for col in cols:
    if col in df.columns:
        df[col] = df[col].fillna("").astype(str)
    else:
        df[col] = ""

df["Actiepunten"] = df[cols].apply(
    lambda row: "; ".join([v.strip() for v in row if v.strip()]),
    axis=1
)

records = []
for _, row in df.iterrows():
    project = row["Projectnummer"]
    actiepunten = row["Actiepunten"]

    # Extra velden (altijd string, nooit NaN)
    projectleider = str(row.get("Projectleider", "") or "").strip()
    type_val = str(row.get("Type", "") or "").strip()

    if actiepunten:
        actiepunten_clean = actiepunten.replace("•", ";")
        for a in actiepunten_clean.split(";"):
            cleaned = a.strip().lstrip("•").strip()
            if cleaned:
                records.append({
                    "Projectnummer": project,
                    "Projectleider": projectleider,
                    "Type": type_val,
                    "Actiepunt": cleaned,
                    "Week": week,
                    "Jaar": jaar
                })

df_new = pd.DataFrame(records)

print("Kolommen in df_new:", df_new.columns.tolist())
print(df_new.head(10).to_string())

# 5. Open of maak werkboek
if os.path.exists(POWERBI_FILE):
    wb = load_workbook(POWERBI_FILE)
else:
    wb = load_workbook(filename=None)

# 6. Haal of maak logsheet
if LOG_SHEET in wb.sheetnames:
    ws = wb[LOG_SHEET]

    # Huidige inhoud ophalen als DataFrame
    existing = pd.DataFrame(ws.values)
    existing.columns = existing.iloc[0]
    existing = existing.drop(0)

    # Filter bestaande data op andere weken dan huidige
    existing = existing[
        ~((existing["Week"] == str(week)) & (existing["Jaar"] == str(jaar)))
    ]

    # Combineer bestaande en nieuwe data
    df_combined = pd.concat([existing, df_new], ignore_index=True)

    # Dubbele regels verwijderen
    df_combined.drop_duplicates(
        subset=["Projectnummer", "Actiepunt", "Week", "Jaar"],
        inplace=True
    )

else:
    wb.create_sheet(LOG_SHEET)
    df_combined = df_new


# 7. Overschrijf logsheet
if LOG_SHEET in wb.sheetnames:
    ws = wb[LOG_SHEET]
    wb.remove(ws)
wb.create_sheet(LOG_SHEET)
ws = wb[LOG_SHEET]

for r in dataframe_to_rows(df_combined, index=False, header=True):
    ws.append(r)

# 8. Opslaan
wb.save(POWERBI_FILE)
print(f"✅ {len(df_new)} actiepunten gelogd voor week {week}, jaar {jaar}.")
