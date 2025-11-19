import os
import re
import html
import pandas as pd
from datetime import datetime
import win32com.client as win32
from collections import defaultdict

# ───────────────────────────────────────────────────────────
OUTPUT_FOLDER   = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Output"
EMAILMAP_FILE   = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Projectleiders.xlsx"
EMAILMAP_SHEET  = "Sheet1"
OVERVIEW_SHEET  = "Overzicht"
TESTMODE        = False  # Alles naar jezelf, maar ontvanger-whitelist blijft verplicht en e-mail moet geldig zijn
SIGNATURE_NAME  = "Bram Gerrits.htm"
REWORD_MAP = {
    "einddatum": "Einddatum verlopen, graag een nieuwe leverdatum doorgeven.",
    "leverdatum": "Leverdatum verlopen, graag een nieuwe leverdatum doorgeven.",
}

# ───────────────────────────────────────────────────────────

def find_latest_overview(folder):
    files = [
        f for f in os.listdir(folder)
        if f.startswith("Overzicht_Projectadministratie_Week") and f.endswith(".xlsx")
    ]
    if not files:
        raise FileNotFoundError(f"Geen Overzicht-bestanden gevonden in: {folder}")
    files.sort(key=lambda x: os.path.getmtime(os.path.join(folder, x)), reverse=True)
    return os.path.join(folder, files[0])

def get_signature_html():
    sig_path = os.path.join(os.environ.get("APPDATA", ""), "Microsoft", "Signatures", SIGNATURE_NAME)
    if os.path.exists(sig_path):
        with open(sig_path, encoding="utf-8") as f:
            return f.read()
    print(f"⚠️ Geen handtekeningbestand gevonden: {sig_path}")
    return "<p>Vriendelijke groet,<br><br>Bram.<br><br> [LET OP: dit is een automatisch gegenereerde mail]</p>"

def to_bool(x) -> bool:
    s = str(x).strip().lower()
    return s in {"true", "1", "ja", "yes", "y", "waar", "ok", "oké", "x"}

def norm_txt(s: str) -> str:
    if not isinstance(s, str):
        return ""
    s = s.replace("\u00A0", " ")  # harde spaties
    s = s.replace("–", "-").replace("—", "-")  # en/em dash -> hyphen
    s = re.sub(r"\s+", " ", s)
    return s.strip().lower()

def is_valid_email(addr: str) -> bool:
    return isinstance(addr, str) and "@" in addr and "." in addr and addr.strip() != ""

# 🔕 Filter regels in PL-mails die (nog) niet gecommuniceerd moeten worden
BLOCKED_PL_PHRASES = [
    "budget kosten",
    "budget opbrengsten",  # dekt ook "budget opbrengsten toevoegen"
]


def filter_actiepunten_tekst(txt: str, blocked=BLOCKED_PL_PHRASES) -> str:
    """
    Verwijdert regels die een van de geblokkeerde (deel)zinnen bevatten (case-insensitive).
    Herschrijft bekende korte codes (Einddatum/Leverdatum verlopen) naar duidelijkere tekst.
    Retourneert de samengevoegde tekst met \n, of '' als alles wegvalt.
    """
    if not isinstance(txt, str) or not txt.strip():
        return ""
    lines = [l.strip() for l in txt.replace("\r\n", "\n").replace("\r", "\n").split("\n")]
    keep = []
    for l in lines:
        low = l.lower()
        # 1) blokkeren op keywords
        if any(b in low for b in blocked):
            continue

        mapped = l
        for key, new in REWORD_MAP.items():
            if key.lower() in low:  # low = l.lower()
                    mapped = f"• {new}"
                    break


        if mapped.strip():
            keep.append(mapped)

    return "\n".join(keep).strip()


# Mapping van actiepunt → ontvangernaam (zoals die in je Projectleiders.xlsx staat)
RECIPIENT_BY_PHRASE = {
    "Gesloten SO met openstaande bestelling": "Inkoop",
    "Gesloten SO met openstaande PO - Prod": "Judith",
    "Gesloten SO met openstaande PO - Proto": "Inkoop",
}
RECIPIENT_BY_PHRASE_NORM = {norm_txt(k): v for k, v in RECIPIENT_BY_PHRASE.items()}

# 1) Laad Overzicht
overview_path = find_latest_overview(OUTPUT_FOLDER)
print(f"📄 Gebruik Overzicht-bestand: {os.path.basename(overview_path)}")

df = pd.read_excel(overview_path, sheet_name=OVERVIEW_SHEET, header=0)
df.columns = df.columns.str.strip()

required_cols = ["Projectnummer", "Projectleider", "Actiepunten Projectleider", "Actiepunten Elders"]
missing = [c for c in required_cols if c not in df.columns]
if missing:
    raise ValueError(f"Ontbrekende kolommen in '{OVERVIEW_SHEET}': {missing}")



# ── Project-niveau Whitelist (kolom in Overzicht) ─────────────────────
# Als kolom "Whitelist" bestaat en NIET leeg is -> project niet mailen
if "Whitelist" in df.columns:
    df["Whitelist_flag"] = df["Whitelist"].notna() & (df["Whitelist"].astype(str).str.strip() != "")
else:
    df["Whitelist_flag"] = False

total_projects = len(df)
total_whitelisted_projects = int(df["Whitelist_flag"].sum())
print(f"📊 Totaal projecten in overzicht: {total_projects}")
print(f"📊 Projecten met Whitelist gevuld (project niet mailen): {total_whitelisted_projects}")

df = df[required_cols + ["Whitelist_flag"]]

# Voor PL-mails: eerst alle regels met Projectleider, daarna filteren op project-whitelist
df_pl_full = df.dropna(subset=["Projectleider"]).copy()
pl_rows_before = len(df_pl_full)
pl_rows_whitelisted = int(df_pl_full["Whitelist_flag"].sum())
df_pl = df_pl_full[~df_pl_full["Whitelist_flag"]].copy()
pl_rows_after = len(df_pl)

print(f"📊 PL-regels totaal: {pl_rows_before}")
print(f"📊 PL-regels met Whitelist (geskipt): {pl_rows_whitelisted}")
print(f"📊 PL-regels na project-whitelist-filter: {pl_rows_after}")

# 2) Laad e-mailmapping (+ ontvanger-whitelist)
df_emails = pd.read_excel(EMAILMAP_FILE, sheet_name=EMAILMAP_SHEET)
df_emails.columns = df_emails.columns.str.strip()
if "Naam" not in df_emails.columns or "Email" not in df_emails.columns:
    raise ValueError("Kolommen 'Naam' en/of 'Email' ontbreken in Projectleiders.xlsx")

df_emails["Naam"]  = df_emails["Naam"].astype(str).str.strip()
df_emails["Email"] = df_emails["Email"].astype(str).str.strip()
# Lege/ongeldige adressen expliciet leeg laten
df_emails.loc[df_emails["Email"].str.lower().isin({"", "nan", "none"}), "Email"] = ""

if "Timestamp" not in df_emails.columns:
        df_emails["Timestamp"] = pd.NaT
else:
        df_emails["Timestamp"] = pd.to_datetime(df_emails["Timestamp"], errors="coerce")


# Ontvanger-niveau whitelist: bepaalt of een persoon überhaupt mails mag ontvangen
if "Whitelist" in df_emails.columns:
    df_emails["Whitelist"] = df_emails["Whitelist"].map(to_bool)
else:
    df_emails["Whitelist"] = False  # streng: alleen expliciet TRUE mag mailen

email_map     = dict(zip(df_emails["Naam"], df_emails["Email"]))
timestamp_map = dict(zip(df_emails["Naam"], df_emails["Timestamp"]))
whitelist_map = dict(zip(df_emails["Naam"], df_emails["Whitelist"]))

def get_email(name: str) -> str:
    addr = (email_map.get(name) or "").strip()
    return addr if is_valid_email(addr) else ""

# 3) Vul elders-bucket (niet filteren op Projectleider, wél op project-whitelist)
elders_bucket = defaultdict(list)  # ontvanger -> list[(Projectnummer, Actiepunt)]
elders_whitelist_skipped = 0

for _, r in df.iterrows():
    # Projecten met Whitelist_flag = True overslaan voor "Elders"
    if r.get("Whitelist_flag", False):
        elders_whitelist_skipped += 1
        continue

    txt_norm = norm_txt(r.get("Actiepunten Elders", ""))
    if not txt_norm:
        continue
    for phrase_norm, recipient in RECIPIENT_BY_PHRASE_NORM.items():
        if phrase_norm in txt_norm:
            # Bewaar de originele, nette phrase voor weergave
            original_phrase = [k for k in RECIPIENT_BY_PHRASE if norm_txt(k) == phrase_norm][0]
            elders_bucket[recipient].append((r["Projectnummer"], original_phrase))

print(f"📊 Elders-regels geskipt door project-whitelist: {elders_whitelist_skipped}")

# 4) Handtekening + Outlook
signature_html = get_signature_html()
try:
    outlook = win32.Dispatch("Outlook.Application")
    _ = outlook.GetNamespace("MAPI")
except Exception as e:
    raise RuntimeError(f"Outlook niet beschikbaar: {e}")

mail_count = 0
week_str = datetime.now().strftime("Week %W, %Y")
test_tag = "[TEST] " if TESTMODE else ""

# 5) Mails: Projectleiders
for leider, sub in df_pl.groupby("Projectleider"):

    # ── Anti-spam check (max 1 mail per dag per persoon) ──
    last_sent = timestamp_map.get(leider)
    today = pd.Timestamp.now().normalize()

    if pd.notna(last_sent) and last_sent.normalize() >= today:
        print(f"⛔ {leider} vandaag al gemaild ({last_sent.date()}) – overslaan")
        continue

    # Ontvanger-whitelist verplicht
    if not whitelist_map.get(leider, False):
        print(f"⛔ Ontvanger niet op whitelist (Projectleiders.xlsx): {leider} – mail overgeslagen")
        continue

    # Ontvanger-whitelist verplicht (óók in TESTMODE)
    if not whitelist_map.get(leider, False):
        print(f"⛔ Ontvanger niet op whitelist (Projectleiders.xlsx): {leider} – mail overgeslagen")
        continue
    # Geldig e-mailadres verplicht
    real_addr = get_email(leider)
    if not real_addr:
        print(f"⛔ Geen geldig e-mailadres voor: {leider} – mail overgeslagen")
        continue

    # Filter lege actiepunten + blokkeerregels
    sub = sub.dropna(subset=["Actiepunten Projectleider"]).copy()
    if sub.empty:
        print(f"ℹ️ Geen actiepunten (na NaN-filter) voor: {leider}")
        continue

    # HTML-tabel opbouwen met filtering
    table_rows = ""
    projecten_in_mail = 0
    for _, row in sub.iterrows():
        filtered_txt = filter_actiepunten_tekst(row.get("Actiepunten Projectleider", ""))
        if not filtered_txt:
            continue  # alle regels geblokkeerd -> project overslaan
        actiepunten_html = html.escape(filtered_txt).replace("\n", "<br>")
        table_rows += (
            f"<tr>"
            f"<td style='text-align:left;'>{row['Projectnummer']}</td>"
            f"<td style='text-align:left;'>{actiepunten_html}</td>"
            f"</tr>"
        )
        projecten_in_mail += 1

    if not table_rows:
        print(f"ℹ️ Na filtering van actiepunten blijft er niets over voor: {leider}")
        continue  # niets te communiceren na filter

    html_body = f"""
    <p>Hallo {str(leider).split()[0]},</p>
    <p>Zou je onderstaande punt(en) kunnen oppakken voor de projecten die aan jouw naam gekoppeld zijn?</p>
    <table border="1" cellpadding="5" cellspacing="0" style="border-collapse: collapse;">
        <tr style="background-color:#4F81BD;color:white;">
            <th style="width: 150px; text-align:left;">Projectnummer</th>
            <th style="width: 300px; text-align:left;">Actiepunten</th>
        </tr>
        {table_rows}
    </table>
    <p>Mocht je nog vragen hebben, laat het gerust weten. <br><br>Alvast bedankt voor je hulp! </p>
    {signature_html}
    """

    to_address = "bram.gerrits@vhe.nl" if TESTMODE else real_addr

    mail = outlook.CreateItem(0)
    mail.To = to_address
    mail.Subject = f"{test_tag}[Projectadministratie] Actiepunten – {str(leider).split()[0]}"
    mail.HTMLBody = html_body
    mail.Display()

    print(f"✅ Mail klaargezet voor: {leider} → {to_address} (projecten in mail: {projecten_in_mail})")
    mail_count += 1
        # Timestamp bijwerken in df_emails (anti-spam)
    today = pd.Timestamp.now().normalize()
    df_emails.loc[df_emails["Naam"] == leider, "Timestamp"] = today



# 6) Mails: Actiepunten Elders (per ontvanger, PER PROJECTNUMMER)
for recipient_name, items in elders_bucket.items():
    if not items:
        continue
    # Ontvanger-whitelist verplicht
    if not whitelist_map.get(recipient_name, False):
        print(f"⛔ Ontvanger (Elders) niet op whitelist: {recipient_name} – mail overgeslagen")
        continue
    # Geldig e-mailadres verplicht
    real_addr = get_email(recipient_name)
    if not real_addr:
        print(f"⛔ Geen geldig e-mailadres voor (Elders): {recipient_name} – mail overgeslagen")
        continue

    # Per projectnummer de lijst met actiepunten (dedup + sort)
    per_project = {}
    for proj, phrase in items:
        per_project.setdefault(proj, [])
        if phrase not in per_project[proj]:
            per_project[proj].append(phrase)

    if not per_project:
        continue

    table_rows = ""
    for proj in sorted(per_project, key=lambda x: str(x)):
        acties_html = "<br>".join(per_project[proj])
        table_rows += (
            f"<tr>"
            f"<td style='text-align:left;'>{proj}</td>"
            f"<td style='text-align:left;'>{acties_html}</td>"
            f"</tr>"
        )

    html_body_elders = f"""
    <p>Hallo {recipient_name},</p>
    <p>Zou je onderstaande punt(en) kunnen oppakken voor de projecten die aan jouw naam gekoppeld zijn?:</p>
    <table border="1" cellpadding="5" cellspacing="0" style="border-collapse: collapse;">
        <tr style="background-color:#4F81BD;color:white;">
            <th style="width: 150px; text-align:left;">Projectnummer</th>
            <th style="width: 300px; text-align:left;">Actiepunten</th>
        </tr>
        {table_rows}
    </table>
    <p>Mocht je nog vragen hebben, laat het gerust weten. <br><br>Alvast bedankt voor je hulp! </p>
    {signature_html}
    """

    to_address = "bram.gerrits@vhe.nl" if TESTMODE else real_addr

    mail = outlook.CreateItem(0)
    mail.To = to_address
    mail.Subject = f"{test_tag}[Projectadministratie] Actiepunten – {recipient_name}"
    mail.HTMLBody = html_body_elders
    mail.Display()

    print(f"✅ Mail (Elders) klaargezet voor: {recipient_name} → {to_address} (projecten in mail: {len(per_project)})")
    mail_count += 1

# Wijzigingen in Timestamp terugschrijven naar Projectleiders.xlsx
with pd.ExcelWriter(EMAILMAP_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    df_emails.to_excel(writer, sheet_name=EMAILMAP_SHEET, index=False)

# ── Timestamp terugschrijven naar Projectleiders.xlsx ──
df_emails.to_excel(EMAILMAP_FILE, sheet_name=EMAILMAP_SHEET, index=False)

# ── Kolombreedtes aanpassen naar 30 ──
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb_pl = load_workbook(EMAILMAP_FILE)
ws_pl = wb_pl[EMAILMAP_SHEET]

for col in range(1, ws_pl.max_column + 1):
    ws_pl.column_dimensions[get_column_letter(col)].width = 30

wb_pl.save(EMAILMAP_FILE)
print("📏 Kolombreedte Projectleiders.xlsx ingesteld op 30.")



print(f"\n📬 Totaal {mail_count} mails klaargezet.")
