import os
import shutil
import pandas as pd
from datetime import date
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ───────────────────────────────────────────────────────────
OUTPUT_FOLDER = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Output"
TEMPLATE_FILE = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Werkbestand Projectadministratie.xlsx"
MAPPING_FILE  = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Kolommenmapping per bron.xlsx"
WORK_ARCHIVE  = os.path.join(OUTPUT_FOLDER, "Werkbestanden archief")
# ───────────────────────────────────────────────────────────


# ————— helpers ——————————————————————————————————————————————

def ensure_dir(path: str):
    if not os.path.exists(path):
        os.makedirs(path)


def move_old_workfiles(output_folder: str, archive_folder: str):
    """Move existing Werkbestand_AlleProjecten_*.xlsx from output → archive."""
    ensure_dir(archive_folder)
    for f in list(os.listdir(output_folder)):
        if f.startswith("Werkbestand_AlleProjecten_") and f.lower().endswith(".xlsx"):
            src = os.path.join(output_folder, f)
            dst = os.path.join(archive_folder, f)
            if os.path.exists(dst):
                base, ext = os.path.splitext(f)
                k = 1
                while os.path.exists(os.path.join(archive_folder, f"{base} ({k}){ext}")):
                    k += 1
                dst = os.path.join(archive_folder, f"{base} ({k}){ext}")
            shutil.move(src, dst)


def find_latest_overview(folder: str) -> str:
    files = [f for f in os.listdir(folder)
             if f.startswith("Overzicht_Projectadministratie_Week") and f.lower().endswith(".xlsx")]
    if not files:
        raise FileNotFoundError(f"No overview file found in {folder}")
    files.sort(key=lambda x: os.path.getmtime(os.path.join(folder, x)), reverse=True)
    return os.path.join(folder, files[0])


def find_col_by_header(ws, header_name: str):
    for cell in ws[2]:
        if cell.value and str(cell.value).strip() == header_name:
            return cell.column
    return None


def move_column_after(ws, col_name_to_move: str, col_name_anchor: str):
    """Moves the column with header col_name_to_move right after col_name_anchor (headers in row 2)."""
    def find_col(header):
        for cell in ws[2]:
            if cell.value and str(cell.value).strip() == header:
                return cell.column
        return None

    src_col = find_col(col_name_to_move)
    anchor_col = find_col(col_name_anchor)
    if not src_col or not anchor_col or src_col == anchor_col + 1:
        return  # missing or already in place

    target_col = anchor_col + 1
    ws.insert_cols(target_col)

    # account for index shift if src_col was at/after target
    src_col_shifted = src_col + (1 if src_col >= target_col else 0)
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        src = ws.cell(row=r, column=src_col_shifted)
        dst = ws.cell(row=r, column=target_col)
        dst.value = src.value
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format

    ws.delete_cols(src_col_shifted)


def rename_header(ws, old: str, new: str):
    """Rename a header in row 2 if present."""
    col = find_col_by_header(ws, old)
    if col:
        ws.cell(row=2, column=col, value=new)


# Robuuste parsing voor ‘Handmatig verwacht resultaat’
POSSIBLE_MANUAL_COLS = [
    "Handmatig verwacht resultaat",
    "Handmatig verwacht resultaat.",
    "Handmatig resultaat",
    "Handmatig",
]

def to_bool(x):
    if isinstance(x, bool):
        return x
    if x is None:
        return False
    if isinstance(x, (int, float)):
        return (not pd.isna(x)) and x != 0
    s = str(x).strip().lower()
    if s in {"true", "waar", "ja", "j", "y", "yes", "1"}:
        return True
    if s in {"false", "onwaar", "nee", "n", "no", "0"}:
        return False
    return False

def get_manual_flag(row: pd.Series) -> bool:
    for col in POSSIBLE_MANUAL_COLS:
        if col in row.index:
            return to_bool(row[col])
    return False


# ————— main ————————————————————————————————————————————————

# 1) Load latest central overview (data source)
overview_path = find_latest_overview(OUTPUT_FOLDER)
df = pd.read_excel(overview_path, header=0, engine="openpyxl")
df.columns = df.columns.str.strip()

# (optioneel: korte log om te zien wat er in de vlag-kolom staat)
if "Handmatig verwacht resultaat" in df.columns:
    try:
        print("🔎 Uniek in 'Handmatig verwacht resultaat':",
              df["Handmatig verwacht resultaat"].astype(str).str.strip().unique().tolist())
    except Exception:
        pass
else:
    print(" Kolom 'Handmatig verwacht resultaat' niet gevonden in centraal bestand.")

# 2) Load mapping and build header map
df_map = pd.read_excel(MAPPING_FILE, header=0)
df_map.rename(columns={df_map.columns[0]: "Header"}, inplace=True)
df_map.columns = df_map.columns.str.strip()

mask_in_out = df_map["Soort"].str.strip().str.lower().isin(["input", "output"])
df_in_out = df_map.loc[mask_in_out]

header_map = {
    str(std).strip(): str(werk).strip()
    for std, werk in zip(df_in_out["Header"], df_in_out["Werkbestand Projectadministratie"])
    if isinstance(werk, str) and str(werk).strip()
}

# Zorg dat onderstaande kolommen sowieso bestaan in header_map
for col in [
    "Projectleider",
    "Omschrijving",
    "Verwacht resultaat",
    "Aangepast resultaat",
    "Algemene informatie",
    "Actiepunten Bram",          # (was Overig)
    "Actiepunten Projectleider",
    "Bespreekpunten",            # (was Actiepunten Bram)
    "Actiepunten Elders",        # ← NIEUW
    "Resultaat gewijzigd",
]:
    header_map.setdefault(col, col)

# 3) Open template workbook
wb = load_workbook(TEMPLATE_FILE)
ws = wb.active

# Remove existing table definitions (if any)
if hasattr(ws, "_tables"):
    ws._tables.clear()

# 4) Ensure 'Aangepast resultaat' exists
template_headers = [str(cell.value).strip() for cell in ws[2] if cell.value is not None]
if "Aangepast resultaat" not in template_headers:
    print("➕ Adding 'Aangepast resultaat' column to template...")
    verw_col = find_col_by_header(ws, "Verwacht resultaat")
    if not verw_col:
        raise ValueError("Kon kolom 'Verwacht resultaat' niet vinden in template (rij 2).")
    insert_at = verw_col + 1
    ws.insert_cols(insert_at)
    ws.cell(row=2, column=insert_at, value="Aangepast resultaat")

# 6) Enforce visual order
move_column_after(ws, "Verwacht resultaat", "Omschrijving")
move_column_after(ws, "Aangepast resultaat", "Verwacht resultaat")
move_column_after(ws, "Actiepunten Bram", "Algemene informatie")  # (was Overig)

# 7) Map the template's visible columns after any moves/renames
template_cols = {}
for cell in ws[2]:
    if cell.value:
        name = str(cell.value).strip()
        if name in header_map.values():
            template_cols[name] = cell.column
if not template_cols:
    raise ValueError("No matching template columns found in row 2")

# Determine bounds (only columns that actually have headers)
header_cols = [cell.column for cell in ws[2] if cell.value is not None]
min_col, max_col = min(header_cols), max(header_cols)

# 8) Styles
hdr_color = "4F81BD"
header_fill = PatternFill(start_color=hdr_color, end_color=hdr_color, fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)
data_font   = Font(size=12)
align_style = Alignment(horizontal="left", vertical="top", wrap_text=True)
even_fill   = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
odd_fill    = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

thin_grey = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Editable inputs (no fill + thin borders)
editable_cols = {"Aangepast resultaat", "Algemene informatie", "Actiepunten Bram"}

# 9) Write data rows (starting at row 3) with manual/auto logic for VR/AR
start_row = 3
df_iter = df.reset_index(drop=True)

# helpers for writing cells
def clear_cell(row_idx_, col_idx_):
    if col_idx_:
        ws.cell(row=row_idx_, column=col_idx_, value=None)

def write_editable_euro(row_idx_, col_idx_, value_):
    cell_ = ws.cell(row=row_idx_, column=col_idx_, value=value_ if pd.notna(value_) else "")
    cell_.fill = PatternFill(fill_type=None)
    cell_.border = thin_grey
    cell_.font = data_font
    cell_.alignment = align_style
    cell_.number_format = "€ #,##0"

def write_normal_euro(row_idx_, col_idx_, value_):
    cell_ = ws.cell(row=row_idx_, column=col_idx_, value=value_)
    cell_.font = data_font
    cell_.alignment = align_style
    cell_.number_format = "€ #,##0"

for row_idx, row in df_iter.iterrows():
    r = start_row + row_idx

    # --- 9a) Specifieke handling Verwacht/Aangepast ---
    vr_col = template_cols.get("Verwacht resultaat")
    ar_col = template_cols.get("Aangepast resultaat")

    vr_val = row.get("Verwacht resultaat")
    ar_src_val = row.get("Aangepast resultaat") if "Aangepast resultaat" in row.index else None
    is_manual = get_manual_flag(row)

    # normaliseer VR voor euro
    if pd.notna(vr_val):
        try:
            vr_val_num = int(round(float(vr_val)))
        except Exception:
            vr_val_num = vr_val
    else:
        vr_val_num = None

    # kies bronwaarde voor AR bij handmatig (gebruik AR-kolom uit centraal als die bestaat)
    manual_val = ar_src_val if (ar_src_val is not None and pd.notna(ar_src_val)) else vr_val_num

    if is_manual and ar_col:
        # handmatig → waarde naar Aangepast, Verwacht leeg
        write_editable_euro(r, ar_col, manual_val)
        clear_cell(r, vr_col)
    else:
        # automatisch → waarde naar Verwacht, Aangepast leeg
        if vr_col:
            write_normal_euro(r, vr_col, vr_val_num)
        clear_cell(r, ar_col)

    # --- 9b) Rest van de kolommen schrijven (skip VR/AR) ---
    for std, werk in header_map.items():
        if werk in {"Verwacht resultaat", "Aangepast resultaat"}:
            continue
        if std in df.columns and werk in template_cols:
            c = template_cols[werk]
            val = row[std]

            # Editable tekstkolommen: geen fill, wel rand
            if werk in {"Algemene informatie", "Actiepunten Bram"}:
                cell = ws.cell(row=r, column=c, value=val if pd.notna(val) else "")
                cell.fill = PatternFill(fill_type=None)
                cell.border = thin_grey
                cell.font = data_font
                cell.alignment = align_style
                continue

            # Bullet-normalisatie
            # Bullet-normalisatie
            if werk in ["Actiepunten Projectleider", "Bespreekpunten", "Actiepunten Elders"] and isinstance(val, str):
                parts = [p.strip() for p in val.replace(";", "\n").split("\n") if p.strip()]
                val = "\n".join(parts)

            cell = ws.cell(row=r, column=c, value=val)
            if werk == "Projectnummer":
                cell.font = Font(size=12, italic=True)
            else:
                cell.font = data_font
            cell.alignment = align_style

            if werk == "Verwacht resultaat" and isinstance(val, (int, float)):
                cell.number_format = "€ #,##0"

    # --- 9c) Rijbanding (niet op editable kolommen) ---
    fill = even_fill if row_idx % 2 == 0 else odd_fill
    for c in range(min_col, max_col + 1):
        col_name = next((name for name, col_idx in template_cols.items() if col_idx == c), None)
        if col_name not in editable_cols:
            ws.cell(row=r, column=c).fill = fill
        else:
            ws.cell(row=r, column=c).border = thin_grey

    ws.row_dimensions[r].height = 50

# 10) Header styling — kleur ALLE headercellen (rij 2)
for c in header_cols:
    cell = ws.cell(row=2, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="left", vertical="center")

# 11) Maak rij 1 ook blauw (zelfde balk als headers)
for c in range(min_col, max_col + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font

# 12) Datum in A1
ws["A1"] = f"Datum: {date.today().strftime('%Y-%m-%d')}"

# 13) Freeze panes
ws.freeze_panes = "A3"

# 14) Column widths
total_rows = start_row + len(df_iter) - 1
for werk, col_idx in template_cols.items():
    col_letter = get_column_letter(col_idx)
    if werk in ["Actiepunten Projectleider", "Bespreekpunten", "Actiepunten Elders"]:
        ws.column_dimensions[col_letter].width = 50
        ws.column_dimensions[col_letter].width = 50
    elif werk in editable_cols:
        ws.column_dimensions[col_letter].width = 30
    else:
        lens = [len(str(ws[f"{col_letter}2"].value or ""))]
        for r in range(start_row, total_rows + 1):
            lens.append(len(str(ws[f"{col_letter}{r}"].value or "")))
        ws.column_dimensions[col_letter].width = max(lens) + 2

# 14b) AutoFilter instellen op rij 2
first_col = get_column_letter(min_col)
last_col  = get_column_letter(max_col)
ws.auto_filter.ref = f"{first_col}2:{last_col}{total_rows}"


# 15) Archive old workfiles, then save new one in Output root
move_old_workfiles(OUTPUT_FOLDER, WORK_ARCHIVE)

today_txt = date.today().strftime("%Y-%m-%d")
out_fn = f"Werkbestand_AlleProjecten_{today_txt}.xlsx"
out_path = os.path.join(OUTPUT_FOLDER, out_fn)
wb.save(out_path)

print(f" Werkbestand gegenereerd: {out_path}")
print(f" Oudere werkbestanden verplaatst naar: {WORK_ARCHIVE}")
print(" Vul waarden in de kolommen zonder kleur (grijze randen).")