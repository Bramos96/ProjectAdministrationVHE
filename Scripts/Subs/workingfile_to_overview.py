import os
import pandas as pd
from openpyxl import load_workbook # allows it to edit databases. cell by cell. 
from openpyxl.utils import get_column_letter #converts numeric column index to letters
from openpyxl.styles import Font # or applying cell formatting 
import time # can be used if time is used. Isn't used. perhaps can be deleted. 

WORKING_FILE = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Werkbestand Projectadministratie.xlsx"
MAPPING_FILE = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Kolommenmapping per bron.xlsx"
OUTPUT_FOLDER = r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Output"

SYNC_COLUMNS = [
    "Algemene informatie",
    "Verwacht resultaat",     # Gets value from "Aangepast resultaat" if not empty
    "Actiepunten Bram",
    "2e Projectleider",
    "Sluiten",
    "Whitelist"
]

MANUAL_INPUT_COLUMN = "Aangepast resultaat"
SHEET_NAME = "Overzicht"
ORANJE_FONT = Font(color="FF6600")
ZWART_FONT = Font(color="000000")
EXTRA_STATUS_COL = "Handmatig verwacht resultaat"

def find_latest_file(folder, prefix):
    files = [f for f in os.listdir(folder)
             if f.startswith(prefix) and f.lower().endswith(".xlsx")]
    if not files:
        raise FileNotFoundError(f"Geen bestand gevonden met prefix '{prefix}' in {folder}")
    files.sort(key=lambda x: os.path.getmtime(os.path.join(folder, x)), reverse=True)
    return os.path.join(folder, files[0])

def is_orange_font(cell):
    if not cell.font or not cell.font.color:
        return False
    try:
        if hasattr(cell.font.color, 'rgb'):
            color = cell.font.color.rgb
            if color:
                return any(o in color.upper() for o in ['FFFF6600', 'FF6600', 'FFF6600', 'FFCC6600'])
        elif hasattr(cell.font.color, 'indexed'):
            return cell.font.color.indexed in [46, 53]
    except Exception:
        return False
    return False

def get_project_row_mapping(worksheet, start_row=2):
    mapping = {}
    for r in range(start_row, worksheet.max_row + 1):
        pn = worksheet[f"A{r}"].value
        if pn:
            mapping[str(pn).strip()] = r
    return mapping

def normalize_sluiten_value(val):
    """Zet alles wat op 'ja' lijkt om naar exact 'Ja'."""
    import pandas as pd

    if pd.isna(val):
        return None

    s = str(val).strip()
    if not s:
        return None

    lower = s.lower()

    # Alles wat begint met 'ja' → Ja (ja, jaa, jaaa, ja hoor, JAA, etc.)
    if lower.startswith("ja"):
        return "Ja"

    # Eventueel nog een paar andere bevestigende varianten
    if lower in {"y", "yes", "1", "true"}:
        return "Ja"

    # Anders gewoon originele (schoongemaakte) tekst teruggeven
    return s


def main():
    print(" Tool 1: Synchroniseer werkbestand centraal bestand")

    try:
        # 1. Laad werkbestand
        print(" Zoek nieuwste werkbestand...")
        workfile_path = find_latest_file(OUTPUT_FOLDER, "Werkbestand_AlleProjecten")
        print(f"   Gevonden: {os.path.basename(workfile_path)}")

        df_work = pd.read_excel(workfile_path, sheet_name="Sheet1", header=1, engine="openpyxl")
        df_work.columns = df_work.columns.str.strip().str.lower()

        # Kolomnamen mappen naar consistente namen
        column_mapping = {
            "project": "projectnummer",
            "actiepunten projectleider": "actiepunten projectleider",  # FIX ✅
            "bespreekpunten": "bespreekpunten",             # FIX ✅
            "actiepunten bram": "actiepunten bram",       # FIX ✅
            "verwacht resultaat": "verwacht resultaat",
            "aangepast resultaat": "aangepast resultaat",
            "algemene informatie": "algemene informatie",
            "2e projectleider": "2e projectleider"

        }
        df_work.rename(columns=column_mapping, inplace=True)

        if "projectnummer" in df_work.columns:
            df_work["projectnummer"] = df_work["projectnummer"].astype(str).str.strip()
            df_work.set_index("projectnummer", inplace=True)
            df_work = df_work[~df_work.index.duplicated(keep="first")]
        else:
            raise Exception("'projectnummer' kolom niet gevonden in werkbestand.")

        wb_work = load_workbook(workfile_path)
        ws_work = wb_work.active
        work_project_to_row = get_project_row_mapping(ws_work, start_row=2)

        # 2. Laad overzichtbestand
        print(" Zoek nieuwste overzichtbestand...")
        overview_path = find_latest_file(OUTPUT_FOLDER, "Overzicht_Projectadministratie_Week")
        print(f"   Gevonden: {os.path.basename(overview_path)}")

        wb = load_workbook(overview_path)
        ws = wb[SHEET_NAME]

        # 4. Zoek kolomletters
        col_map = {}
        header_row = list(ws[1])
        for cell in header_row:
            if cell.value and cell.value.strip() in SYNC_COLUMNS:
                col_map[cell.value.strip()] = get_column_letter(cell.column)

        if EXTRA_STATUS_COL not in [c.value for c in header_row if c.value]:
            print(f" Voeg statuskolom toe: {EXTRA_STATUS_COL}")
            ws.cell(row=1, column=ws.max_column+1, value=EXTRA_STATUS_COL)
            col_map[EXTRA_STATUS_COL] = get_column_letter(ws.max_column)
        else:
            for cell in header_row:
                if cell.value and cell.value.strip() == EXTRA_STATUS_COL:
                    col_map[EXTRA_STATUS_COL] = get_column_letter(cell.column)

                # --- 4b. Zorg dat alle SYNC_COLUMNS als header bestaan; voeg ontbrekende toe en map ze ---
        existing_headers = [c.value.strip() for c in ws[1] if c.value]
        for needed in SYNC_COLUMNS:
            if needed not in existing_headers:
                ws.cell(row=1, column=ws.max_column + 1, value=needed)
                col_map[needed] = get_column_letter(ws.max_column)
            else:
                if needed not in col_map:  # als header er wel was maar nog niet gemapt
                    for cell in ws[1]:
                        if cell.value and cell.value.strip() == needed:
                            col_map[needed] = get_column_letter(cell.column)


        # 5. Synchroniseer
        updates = 0
        manual_detected = 0

        print(" Synchroniseer data...")
        for row_idx in range(2, ws.max_row + 1):
            projectnummer = ws[f"A{row_idx}"].value
            if projectnummer:
                projectnummer = str(projectnummer).strip()
                if projectnummer in df_work.index:

                    for kolom in SYNC_COLUMNS:
                        kolom_lower = kolom.lower()
                        if kolom_lower in df_work.columns:
                            cell_ref = f"{col_map[kolom]}{row_idx}"
                            new_val = df_work.at[projectnummer, kolom_lower]
                            old_val = ws[cell_ref].value

                            # -------------------------------------------------
                            # 1) Speciaal geval: Verwacht resultaat + overrides
                            # -------------------------------------------------
                            if kolom == "Verwacht resultaat":
                                is_manual = False
                                final_value = new_val

                                manual_col_lower = MANUAL_INPUT_COLUMN.lower()
                                if manual_col_lower in df_work.columns:
                                    manual_override = df_work.at[projectnummer, manual_col_lower]
                                    if pd.notna(manual_override) and str(manual_override).strip() != "":
                                        final_value = manual_override
                                        is_manual = True
                                        print(f"        HANDMATIG: {projectnummer} = €{final_value}")

                                if pd.notna(final_value) and str(final_value).strip() != str(old_val).strip():
                                    ws[cell_ref].value = final_value
                                    updates += 1

                                status_col_letter = col_map[EXTRA_STATUS_COL]
                                if is_manual:
                                    ws[cell_ref].font = ORANJE_FONT
                                    ws[f"{status_col_letter}{row_idx}"].value = True
                                    manual_detected += 1
                                else:
                                    ws[cell_ref].font = ZWART_FONT
                                    ws[f"{status_col_letter}{row_idx}"].value = False

                            # -------------------------------------------------
                            # 2) Speciaal geval: Sluiten → normaliseren naar "Ja"
                            # -------------------------------------------------
                            elif kolom == "Sluiten":
                                normalized = normalize_sluiten_value(new_val)
                                if normalized is not None and str(normalized).strip() != str(old_val).strip():
                                    ws[cell_ref].value = normalized
                                    updates += 1

                            # -------------------------------------------------
                            # 3) Standaard synchronisatie
                            # -------------------------------------------------
                                                        # -------------------------------------------------
                            # 3) Standaard synchronisatie  (ook leegmaken toestaan)
                            # -------------------------------------------------
                            else:
                                # Normaliseer new_val
                                if pd.isna(new_val):
                                    new_clean = ""
                                else:
                                    new_clean = str(new_val).strip()

                                # Normaliseer old_val
                                if old_val is None:
                                    old_clean = ""
                                else:
                                    old_clean = str(old_val).strip()

                                # Alleen updaten als er verschil is
                                if new_clean != old_clean:
                                    if new_clean == "":
                                        # Leegmaken van waarde
                                        ws[cell_ref].value = None
                                    else:
                                        # Normale update
                                        ws[cell_ref].value = new_val
                                    updates += 1


        print(" Sla overzichtbestand op...")
        wb.save(overview_path)

        print(f" Synchronisatie voltooid!")
        print(f"    Gewijzigde waarden: {updates}")
        print(f"    Manual overrides: {manual_detected}")

    except Exception as e:
        print(f" FOUT: {e}")
        raise

if __name__ == "__main__":
    main()
