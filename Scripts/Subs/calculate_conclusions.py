import os
import pandas as pd
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# ───────────────────────────────────────────
# PAD NAAR OUTPUT
# ───────────────────────────────────────────
OUTPUT_FILE = (
    r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration"
    r"\Output\Overzicht_Projectadministratie_Week{week}_{year}.xlsx"
)

# ───────────────────────────────────────────
# ACTIEPUNTEN PROJECTLEIDER
# ───────────────────────────────────────────
def is_closed_verkooporder(row) -> bool:
    """Verkooporder wordt als gesloten gezien als:
    - Sluiten == 'Ja', of
    - bij één van de kolommen Openstaande bestelling / SO / PO iets ingevuld is
      (Ja, Nee, wat dan ook).
    """
    sluiten = str(row.get("Sluiten", "")).strip().lower()
    if sluiten == "ja":
        return True

    for col in ["Openstaande bestelling", "Openstaande SO", "Openstaande PO"]:
        val = row.get(col, "")
        if pd.notna(val) and str(val).strip() != "":
            return True

    return False



def make_actions_projectleider(row, today):
    # Als verkooporder gesloten is → geen actiepunten meer voor PL
    if is_closed_verkooporder(row):
        return ""

    bullets = []
    ed_ts = pd.to_datetime(row.get("Einddatum"), errors="coerce")
    if pd.notna(ed_ts) and today > ed_ts.date():
        bullets.append("• Einddatum verlopen")

    lv_ts = pd.to_datetime(row.get("Leverdatum"), errors="coerce")
    if pd.notna(lv_ts) and today > lv_ts.date():
        bullets.append("• Leverdatum(s) verlopen")

    bk = row.get("Budget Kosten")
    try:
        if pd.notna(bk) and float(bk) == 0:
            bullets.append("• Budget kosten toevoegen")
    except:
        pass

    bo = row.get("Budget Opbrengsten")
    try:
        if pd.notna(bo) and float(bo) == 0:
            bullets.append("• Budget opbrengsten toevoegen")
    except:
        pass

    return "\n".join(bullets)



# ───────────────────────────────────────────
# BESPREekpunten (ALLEEN deze twee!)
# ───────────────────────────────────────────
def make_bespreekpunten(row):
    bullets = []

    def to_float(x):
        try:
            return float(x)
        except:
            return None

    def is_yes(v):
        return str(v).strip().lower() in {"ja", "true", "1", "y", "yes"}

    # 1) Negatief resultaat bespreken
    vr = to_float(row.get("Verwacht resultaat"))
    if vr is not None and vr < 0:
        bullets.append("• Negatief resultaat bespreken")

    # 2) Dochterproject-indicator
    val_so = row.get("Openstaande SO", "")
    if is_yes(val_so):
        bullets.append("• Gesloten SO, maar openstaande SO dochterproject")

    # 3) Kunnen we sluiten?
    warning_txt = str(row.get("Warning", "")).strip()

    cols_sc = ["Openstaande bestelling", "Openstaande SO", "Openstaande PO"]
    vals_sc = []
    for c in cols_sc:
        v = row.get(c)
        if pd.isna(v):
            vals_sc.append("")          # echte leegtes → echt leeg
        else:
            vals_sc.append(str(v).strip())

    any_filled = any(v != "" for v in vals_sc)

    if warning_txt == "" and any_filled:
        bullets.append("• Kunnen we deze sluiten?")

    return "\n".join(bullets)




# ───────────────────────────────────────────
# INFORMATIE (opbrengsten binnen + gesloten SO definitief)
# ───────────────────────────────────────────
# ───────────────────────────────────────────
# INFORMATIE (alleen gesloten SO definitief)
# ───────────────────────────────────────────
def make_informatie(row):
    bullets = []

    def is_nee(v):
        s = str(v).strip().lower()
        return s in {"nee", "no", "n", "false", "0"}

    # 1) Gesloten SO → project sluiten na goedkeuring
    cols_sc = ["Openstaande bestelling", "Openstaande SO", "Openstaande PO"]
    vals = [row.get(c, "") for c in cols_sc]

    all_three_nee = (
        all(is_nee(v) for v in vals)
        and all(str(v).strip() != "" for v in vals)
    )

    if all_three_nee:
        bullets.append("• Gesloten SO, project sluiten na goedkeuring")

    return "\n".join(bullets)

# ───────────────────────────────────────────
# WARNING-KOLOM (opbrengsten mismatch)
# ───────────────────────────────────────────
def make_warning(row):
    def to_float(x):
        try:
            return float(x)
        except:
            return None

    bo = to_float(row.get("Budget Opbrengsten"))
    wo = to_float(row.get("Werkelijke opbrengsten"))

    # Als één van de twee ontbreekt → geen warning
    if bo is None or wo is None:
        return ""

    # Als het niet gelijk is → warning
    if bo != wo:
        return "Opbrengsten niet in orde"

    # Exact gelijk → geen melding
    return ""


# ───────────────────────────────────────────
# PROTO/PROD
# ───────────────────────────────────────────
def make_proto_prod(row):
    t = str(row.get("Type", "")).strip().lower()

    mapping = {
        "former qnq customers": "Proto",
        "orders handel": "Proto",
        "orders kabelafdeling": "Prod",
        "orders kastenbouw asml": "Proto",
        "orders projecten": "Proto",
        "orders xt sets": "Prod",
        "proto": "Proto",
        "service orders": "Proto",
    }

    return mapping.get(t, "")

# ───────────────────────────────────────────
# TIERS OP BASIS VAN 5 CHECKS
# ───────────────────────────────────────────

def _tier_checks_count(row) -> int:
    """Geef terug hoeveel van de 4 checks zijn behaald."""

    # Check 1: Verkooporder gesloten indicator
    cols_closed = ["Openstaande bestelling", "Openstaande SO", "Openstaande PO"]
    closed_indicator = False
    for c in cols_closed:
        v = row.get(c)
        if pd.isna(v):
            continue
        s = str(v).strip()
        if s != "":
            closed_indicator = True
            break

    # Check 2: Warning moet leeg zijn (geen tekst)
    v_warning = row.get("Warning", "")
    if pd.isna(v_warning):
        warning_empty = True
    else:
        warning_empty = str(v_warning).strip() == ""

    # Check 3: Sluiten moet 'ja' zijn
    sluiten_val = str(row.get("Sluiten", "")).strip().lower()
    sluiten_ok = (sluiten_val == "ja")

    # Check 4: Actiepunten Bram moet leeg zijn
    v_bram = row.get("Actiepunten Bram", "")
    if pd.isna(v_bram):
        bram_empty = True
    else:
        bram_empty = str(v_bram).strip() == ""

    checks = [closed_indicator, warning_empty, sluiten_ok, bram_empty]
    return sum(1 for ok in checks if ok)


def make_tier1(row):
    return _tier_checks_count(row) == 4

def make_tier2(row):
    return _tier_checks_count(row) == 3

def make_tier3(row):
    return _tier_checks_count(row) == 2

def make_tier4(row):
    return _tier_checks_count(row) == 1

def make_tier5(row):
    return _tier_checks_count(row) == 0


# ───────────────────────────────────────────
# ACTIEPUNTEN ELDERS
# ───────────────────────────────────────────
def make_actiepunten_elders(row):
    val_best = str(row.get("Openstaande bestelling", "")).strip().lower()
    val_po = str(row.get("Openstaande PO", "")).strip().lower()

    yes_best = val_best in {"ja", "true", "1", "y", "yes"}
    yes_po = val_po in {"ja", "true", "1", "y", "yes"}

    t = str(row.get("Type", "")).strip().lower()
    proto_types = {"orders handel", "orders projecten", "service orders"}
    prod_types  = {"orders kabelafdeling", "orders kastenbouw asml"}

    bullets = []

    if yes_best:
        bullets.append("• Gesloten SO met openstaande bestelling")

    if yes_po:
        if t in proto_types:
            bullets.append("• Gesloten SO met openstaande PO - Proto")
        elif t in prod_types:
            bullets.append("• Gesloten SO met openstaande PO - Prod")

    ntr = row.get("Niet toegewezen regels")
    if pd.notna(ntr):
        s = str(ntr).strip()
        if s not in {"", "0", "0.0"}:
            bullets.append("• Orderregel(s) toewijzen aan PR")

    return "\n".join(bullets)


# ───────────────────────────────────────────
# MAIN
# ───────────────────────────────────────────
def main():

    today = date.today()
    week = today.isocalendar()[1]
    year = today.year

    from pathlib import Path
    files = list(Path(r"C:\Users\bram.gerrits\Desktop\Automations\ProjectAdministration\Output").glob(
        "Overzicht_Projectadministratie_Week*.xlsx"
    ))
    if not files:
        raise FileNotFoundError("Geen centrale overzichten gevonden.")
    path = max(files, key=lambda f: f.stat().st_mtime)
    print(f"\n Gekozen bestand: {path.name}\n")

    df = pd.read_excel(path, header=0, engine="openpyxl")

    # BEREKEN KOLLOMMEN
    df["Actiepunten Projectleider"] = df.apply(lambda r: make_actions_projectleider(r, today), axis=1)
    df["Informatie"] = df.apply(make_informatie, axis=1)
    df["Warning"] = df.apply(make_warning, axis=1)
    df["Bespreekpunten"] = df.apply(make_bespreekpunten, axis=1)
    df["Actiepunten Elders"] = df.apply(make_actiepunten_elders, axis=1)
    df["Proto/Prod"] = df.apply(make_proto_prod, axis=1)
    df["Tier 1"] = df.apply(make_tier1, axis=1)
    df["Tier 2"] = df.apply(make_tier2, axis=1)
    df["Tier 3"] = df.apply(make_tier3, axis=1)
    df["Tier 4"] = df.apply(make_tier4, axis=1)
    df["Tier 5"] = df.apply(make_tier5, axis=1)

    # Convert booleans to 0/1 for Power BI compatibility
    for col in ["Tier 1", "Tier 2", "Tier 3", "Tier 4", "Tier 5"]:
        df[col] = df[col].astype(int)




    wb = load_workbook(path)
    ws = wb["Overzicht"]

    # HEADERS
    col_letters = {}
    for cell in ws[1]:
        if cell.value:
            col_letters[cell.value.strip()] = cell.column_letter

    # Voeg Informatie toe als header niet bestaat
    if "Informatie" not in col_letters:
        last = ws.max_column + 1
        ws.cell(row=1, column=last).value = "Informatie"
        col_letters["Informatie"] = ws.cell(row=1, column=last).column_letter
        ws.column_dimensions[col_letters["Informatie"]].width = 50

    # Voeg Proto/Prod toe indien nodig
    if "Proto/Prod" not in col_letters:
        last = ws.max_column + 1
        ws.cell(row=1, column=last).value = "Proto/Prod"
        col_letters["Proto/Prod"] = ws.cell(row=1, column=last).column_letter
        ws.column_dimensions[col_letters["Proto/Prod"]].width = 12

        # Voeg Warning toe indien nodig
    if "Warning" not in col_letters:
        last = ws.max_column + 1
        ws.cell(row=1, column=last).value = "Warning"
        col_letters["Warning"] = ws.cell(row=1, column=last).column_letter
        ws.column_dimensions[col_letters["Warning"]].width = 30
    
    # Voeg Tier 1 toe indien nodig
    if "Tier 1" not in col_letters:
        last = ws.max_column + 1
        ws.cell(row=1, column=last).value = "Tier 1"
        col_letters["Tier 1"] = ws.cell(row=1, column=last).column_letter
        ws.column_dimensions[col_letters["Tier 1"]].width = 12

    # Voeg Tier 2 toe indien nodig
    if "Tier 2" not in col_letters:
        last = ws.max_column + 1
        ws.cell(row=1, column=last).value = "Tier 2"
        col_letters["Tier 2"] = ws.cell(row=1, column=last).column_letter
        ws.column_dimensions[col_letters["Tier 2"]].width = 12

    # Voeg Tier 3 toe indien nodig
    if "Tier 3" not in col_letters:
        last = ws.max_column + 1
        ws.cell(row=1, column=last).value = "Tier 3"
        col_letters["Tier 3"] = ws.cell(row=1, column=last).column_letter
        ws.column_dimensions[col_letters["Tier 3"]].width = 12

    # Voeg Tier 4 toe indien nodig
    if "Tier 4" not in col_letters:
        last = ws.max_column + 1
        ws.cell(row=1, column=last).value = "Tier 4"
        col_letters["Tier 4"] = ws.cell(row=1, column=last).column_letter
        ws.column_dimensions[col_letters["Tier 4"]].width = 12

# Voeg Tier 5 toe indien nodig
    if "Tier 5" not in col_letters:
        last = ws.max_column + 1
        ws.cell(row=1, column=last).value = "Tier 5"
        col_letters["Tier 5"] = ws.cell(row=1, column=last).column_letter
        ws.column_dimensions[col_letters["Tier 5"]].width = 12

    # WEGSCHRIJVEN VAN ALLE KOLLOMMEN
    def write_column(name):
        if name in col_letters:
            col = col_letters[name]
            for idx, val in enumerate(df[name], start=2):

                # Voor Tier-kolommen: True/False → 1/0
                if name.startswith("Tier "):
                    if pd.isna(val):
                        cell_value = None
                    else:
                        cell_value = int(bool(val))  # True -> 1, False -> 0
                else:
                    cell_value = val  # alles zoals het was

                ws[f"{col}{idx}"].value = cell_value
                ws[f"{col}{idx}"].alignment = Alignment(wrap_text=True)


    write_column("Actiepunten Projectleider")
    write_column("Bespreekpunten")
    write_column("Informatie")
    write_column("Warning") 
    write_column("Actiepunten Elders")
    write_column("Proto/Prod")
    write_column("Tier 1")
    write_column("Tier 2")
    write_column("Tier 3")
    write_column("Tier 4")
    write_column("Tier 5")



    # ZET INFORMATIE OP BREEDTE 50
    ws.column_dimensions[col_letters["Informatie"]].width = 50
    ws.column_dimensions[col_letters["Warning"]].width = 50

    wb.save(path)
    print(f"\n  Script afgerond — gegevens opgeslagen in: {os.path.basename(path)}\n")


if __name__ == "__main__":
    main()
